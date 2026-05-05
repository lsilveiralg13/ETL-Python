import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def aplicar_formatacao_belmicro(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    fonte_corpo = Font(name='Calibri', size=9)
    fonte_cabecalho = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
    alinhamento_central = Alignment(horizontal='center', vertical='center')
    preenchimento_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = fonte_corpo
            cell.alignment = alinhamento_central

    for cell in ws[1]:
        cell.font = fonte_cabecalho
        cell.fill = preenchimento_preto

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 4

    wb.save(caminho_arquivo)

def consolidar_paradas_2025():
    diretorio = r'C:\Users\lucas.barros\OneDrive - BELMICRO TECNOLOGIA SA\Área de Trabalho\Scripts Python'
    arquivo_origem = os.path.join(diretorio, 'Registro parada na Linha 2025.xlsx')
    arquivo_saida = os.path.join(diretorio, 'FATO_CAPACIDADE_PRODUTIVA.xlsx')

    abas_alvo = [
        "JANEIRO25", "FEVEREIRO25", "MARCO25", "ABRIL25", "MAIO25", "JUNHO25",
        "JULHO25", "AGOSTO25", "SETEMBRO25", "OUTUBRO25", "NOVEMBRO25", "DEZEMBRO25"
    ]
    
    # Colunas que queremos, mas agora trataremos variações
    colunas_desejadas = ["Data", "Hora", "Linha", "Capacidade Produtiva Média", "Tipo produção"]
    
    lista_dfs = []

    if not os.path.exists(arquivo_origem):
        print(f"❌ Erro: Arquivo não encontrado em {arquivo_origem}")
        return

    try:
        xls = pd.ExcelFile(arquivo_origem)
        # Limpa espaços de todos os nomes de abas existentes na planilha
        abas_reais = {nome.strip().upper(): nome for nome in xls.sheet_names}

        for aba_procurada in abas_alvo:
            if aba_procurada in abas_reais:
                nome_real_aba = abas_reais[aba_procurada]
                print(f"-> Extraindo dados da aba: {nome_real_aba}")
                
                df_aba = pd.read_excel(xls, sheet_name=nome_real_aba)
                
                # Normalização das colunas: remove espaços e coloca em maiúsculo para comparar
                df_aba.columns = [str(c).strip() for c in df_aba.columns]
                
                # Mapeamento inteligente para a coluna de Capacidade (ajuda com acentos/espaços)
                mapeamento = {}
                for col in df_aba.columns:
                    col_upper = col.upper()
                    if "CAPACIDADE" in col_upper and "M" in col_upper:
                        mapeamento[col] = "Capacidade Produtiva Média"
                    elif "DATA" in col_upper: mapeamento[col] = "Data"
                    elif "HORA" in col_upper: mapeamento[col] = "Hora"
                    elif "LINHA" in col_upper: mapeamento[col] = "Linha"
                    elif "TIPO" in col_upper: mapeamento[col] = "Tipo produção"

                df_aba = df_aba.rename(columns=mapeamento)
                
                # Filtra apenas o que mapeamos e que realmente existe
                cols_finais = [c for c in colunas_desejadas if c in df_aba.columns]
                df_filtrado = df_aba[cols_finais].copy()
                df_filtrado['Mes_Referencia'] = aba_procurada
                
                lista_dfs.append(df_filtrado)
            else:
                print(f"⚠️ Aba {aba_procurada} não encontrada. Verifique se há espaços extras no Excel.")

        if lista_dfs:
            df_final = pd.concat(lista_dfs, ignore_index=True)
            df_final.to_excel(arquivo_saida, index=False)
            aplicar_formatacao_belmicro(arquivo_saida)
            print(f"\n✅ SUCESSO! Arquivo salvo em: {arquivo_saida}")
        else:
            print("❌ Nenhuma aba foi processada com sucesso.")

    except Exception as e:
        print(f"❌ Erro: {e}")

if __name__ == "__main__":
    consolidar_paradas_2025()