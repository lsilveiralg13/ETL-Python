import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Font, Alignment
import time

# ==========================================================
# 🔧 CONFIGURAÇÕES E COLUNAS CHAVE
# ==========================================================

COLUNAS_CHAVE = {
    'COD_PRODUTO': 'Cód. Produto',
    'REFERENCIA': 'Referência',
    'DESCRICAO': 'Descrição Produto',
    'CATEGORIA': 'GRUPO',
    'SUBGRUPO': 'SUBGRUPO',
    'COD_UF': 'COD. UF',
    'CHAVE_MMM': 'CHAVE_MMM',
    'QUANTIDADE': 'Quantidade',
    'VALOR': 'Faturado',
    'ESTOQUE': 'Estoque',
    'ESTADO_IS': 'ESTADO/IS',
    'REGIAO': 'REGIÃO',
}

DIRETORIO_BASE = os.path.abspath(r'C:\Users\lucasbarros\OneDrive - CTC FRANCHISING S A\Área de Trabalho\Scripts Python')
ARQUIVO_BASE_SANKHYA = os.path.join(DIRETORIO_BASE, 'Base de Produtos Mais Vendidos - Showroom.xlsx')
DIRETORIO_IMAGENS = os.path.join(DIRETORIO_BASE, 'IMAGENS')

SUBGRUPOS_PARA_PERGUNTAR = ['CHINELO', 'SACOLAS', 'REVENDA', 'MARY KAY']

# ==========================================================
# 🚀 FUNÇÃO PRINCIPAL
# ==========================================================

def consolidar_relatorio(filtros: dict):

    # 1) Ler a base
    try:
        df_bruto = pd.read_excel(ARQUIVO_BASE_SANKHYA, sheet_name='BASE PRODUTOS')
        print(f"Base Sankhya lida com sucesso. Total de {len(df_bruto)} linhas.")
    except Exception as e:
        print(f"❌ ERRO ao ler planilha: {e}")
        return

    df_bruto.columns = df_bruto.columns.str.strip()

    # 2) Renomear colunas
    mapa_renomeacao = {
        COLUNAS_CHAVE['COD_PRODUTO']: 'CODIGO_SANKHYA',
        COLUNAS_CHAVE['REFERENCIA']: 'REFERENCIA_PRODUTO',
        COLUNAS_CHAVE['DESCRICAO']: 'DESCRICAO_PRODUTO',
        COLUNAS_CHAVE['CATEGORIA']: 'CATEGORIA',
        COLUNAS_CHAVE['SUBGRUPO']: 'SUBGRUPO',
        COLUNAS_CHAVE['COD_UF']: 'COD_UF',
        COLUNAS_CHAVE['CHAVE_MMM']: 'CHAVE_MMM',
        COLUNAS_CHAVE['QUANTIDADE']: 'QTD_VENDIDA',
        COLUNAS_CHAVE['VALOR']: 'VALOR_TOTAL_VENDA',
        COLUNAS_CHAVE['ESTADO_IS']: 'INSIDE_SALES',
        COLUNAS_CHAVE['REGIAO']: 'REGIAO',
        COLUNAS_CHAVE['ESTOQUE']: 'ESTOQUE'
    }
    df_bruto.rename(columns=mapa_renomeacao, inplace=True)

    # 2.1) Fallback de posição
    if 'REGIAO' not in df_bruto.columns and len(df_bruto.columns) >= 20:
        df_bruto.rename(columns={df_bruto.columns[19]: 'REGIAO'}, inplace=True)
    if 'INSIDE_SALES' not in df_bruto.columns and len(df_bruto.columns) >= 21:
        df_bruto.rename(columns={df_bruto.columns[20]: 'INSIDE_SALES'}, inplace=True)

    # 3) Padronização inicial
    for col in ['CATEGORIA', 'SUBGRUPO', 'COD_UF', 'REGIAO', 'INSIDE_SALES', 'CHAVE_MMM']:
        if col in df_bruto.columns:
            df_bruto[col] = df_bruto[col].astype(str).str.upper().str.strip()

    if 'CODIGO_SANKHYA' in df_bruto.columns:
        df_bruto['CODIGO_SANKHYA'] = df_bruto['CODIGO_SANKHYA'].astype(str).str.strip()

    for col in ['QTD_VENDIDA', 'VALOR_TOTAL_VENDA']:
        if col in df_bruto.columns:
            df_bruto[col] = pd.to_numeric(df_bruto[col], errors='coerce').fillna(0)

    # 4) EXCLUSÃO INTERATIVA DE SUBGRUPOS
    if 'SUBGRUPO' in df_bruto.columns:
        print("\n--- VALIDAÇÃO DE EXCLUSÃO DE SUBGRUPOS ---")
        subgrupos_a_excluir = []
        for sg in SUBGRUPOS_PARA_PERGUNTAR:
            resp = input(f"Deseja EXCLUIR o subgrupo '{sg}' do relatório? (S/N): ").strip().upper()
            if resp == 'S':
                subgrupos_a_excluir.append(sg)
        
        if subgrupos_a_excluir:
            antes = len(df_bruto)
            df_bruto = df_bruto[~df_bruto['SUBGRUPO'].isin(subgrupos_a_excluir)]
            print(f"→ Aplicado: {antes - len(df_bruto)} linhas removidas ({', '.join(subgrupos_a_excluir)}).")

    # 5) Filtros interativos
    df_filtrado = df_bruto.copy()
    for coluna, valor in filtros.items():
        if valor and coluna in df_filtrado.columns:
            v = str(valor).upper().strip()
            print(f"→ Filtro: {coluna} = {v}")
            df_filtrado = df_filtrado[df_filtrado[coluna] == v]

    if df_filtrado.empty:
        print("⚠️ Nenhum dado encontrado após aplicar filtros.")
        return

    # 5.1) Filtro de estoque
    opcao_estoque = input("\nDeseja excluir produtos sem estoque? (S/N): ").strip().upper()
    if opcao_estoque == "S":
        if 'ESTOQUE' in df_filtrado.columns:
            antes = len(df_filtrado)
            df_filtrado = df_filtrado[df_filtrado['ESTOQUE'] > 0]
            print(f"→ Estoque aplicado: {antes - len(df_filtrado)} produtos removidos.")

    # 6) Consolidação e ranking
    df_agrup = df_filtrado.groupby(
        ['REFERENCIA_PRODUTO', 'CODIGO_SANKHYA', 'DESCRICAO_PRODUTO', 'CATEGORIA']
    ).agg(VENDAS_TOTAIS=('QTD_VENDIDA', 'sum')).reset_index()

    df_agrup['RANK_CATEGORIA'] = df_agrup.groupby('CATEGORIA')['VENDAS_TOTAIS'].rank(
        method='dense', ascending=False
    ).astype(int)

    df_top10 = (
        df_agrup[df_agrup['RANK_CATEGORIA'] <= 40]
        .sort_values(by=['CATEGORIA', 'RANK_CATEGORIA', 'VENDAS_TOTAIS'], ascending=[True, True, False])
        .groupby('CATEGORIA')
        .head(40)
        .reset_index(drop=True)
    )

    exportar_e_formatar(df_top10, filtros)

# ==========================================================
# 🎨 EXPORTAÇÃO E FORMATAÇÃO (ESTÉTICA COMPLETA)
# ==========================================================

def exportar_e_formatar(df_dados, filtros):
    uf = filtros.get('COD_UF', '') or filtros.get('REGIAO', '') or filtros.get('INSIDE_SALES', 'GERAL')
    cat = filtros.get('CATEGORIA', 'TODAS')
    nome_arquivo = f"Relatorio_Top10_{uf}_{cat}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho_saida = os.path.join(DIRETORIO_BASE, nome_arquivo)

    df_export = df_dados.copy()
    df_export['Rótulos de Linha'] = df_export['REFERENCIA_PRODUTO']
    df_export['Cod. Sankhya']    = df_export['CODIGO_SANKHYA']
    df_export['Descrição']       = df_export['DESCRICAO_PRODUTO']
    df_export['Ranking']         = df_export['RANK_CATEGORIA']
    df_export['VENDAS_TOTAIS']   = df_export['VENDAS_TOTAIS']

    colunas_finais = ['Rótulos de Linha', 'Cod. Sankhya', 'Descrição', 'Ranking', 'VENDAS_TOTAIS']
    df_export = df_export[colunas_finais]

    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Top_10_Consolidado', index=False, startrow=1)

    try:
        wb = load_workbook(caminho_saida)
        ws = wb['Top_10_Consolidado']
        ws.views.sheetView[0].showGridlines = False

        style_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        font_titulo = Font(color="FFFFFFFF", bold=True, size=14)
        cor_preta = '00000000'

        ws.insert_cols(4)
        ws.cell(row=2, column=4, value="Foto").alignment = style_center

        # Título Formatado
        ws.merge_cells('A1:F1')
        ws['A1'].value = f"TOP 10 - {cat.upper()} MAIS VENDIDOS - {uf.upper()}"
        ws['A1'].fill = PatternFill(start_color=cor_preta, end_color=cor_preta, fill_type="solid")
        ws['A1'].font = font_titulo
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Larguras e Alturas
        larguras = {'A': 15.09, 'B': 15.09, 'C': 42.09, 'D': 20, 'E': 15.09, 'F': 15.09}
        for col, width in larguras.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 19
        ws.row_dimensions[2].height = 19
        for row_num in range(3, len(df_dados) + 3):
            ws.row_dimensions[row_num].height = 50

        # Alinhamento Geral
        for row in ws.iter_rows(min_row=2, max_row=len(df_dados)+2, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = style_center

        print("\n→ Inserindo imagens no Excel...")
        for index, row in df_dados.iterrows():
            codigo = str(row['CODIGO_SANKHYA']).strip()
            img_path = None
            for ext in ['.jpg', '.png', '.jpeg']:
                p = os.path.join(DIRETORIO_IMAGENS, f"{codigo}{ext}")
                if os.path.exists(p):
                    img_path = p
                    break

            row_excel = index + 3
            if img_path:
                try:
                    img = OpenpyxlImage(img_path)
                    img.width, img.height = 80, 80
                    ws.add_image(img, f'D{row_excel}')
                except Exception:
                    ws.cell(row=row_excel, column=4, value="Erro Img")
            else:
                ws.cell(row=row_excel, column=4, value="S/ Img")

        wb.save(caminho_saida)
        print(f"[SUCESSO] Relatório: {caminho_saida}")

    except Exception as e:
        print(f"❌ Erro na formatação: {e}")

if __name__ == "__main__":
    os.makedirs(DIRETORIO_IMAGENS, exist_ok=True)
    print("\n1-UF | 2-Região | 3-Inside Sales")
    opcao = input("Opção: ").strip()
    filtros = {}
    if opcao == "1":
        filtros['COD_UF'] = input("Informe UF: ").strip()
    elif opcao == "2":
        filtros['REGIAO'] = input("Informe Região: ").strip()
    elif opcao == "3":
        filtros['INSIDE_SALES'] = input("Informe Inside Sales: ").strip()
    
    filtros['CATEGORIA'] = input("Categoria: ").strip()
    consolidar_relatorio(filtros)
    