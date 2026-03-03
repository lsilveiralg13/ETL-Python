import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
import time

# ==========================================================
# 🔧 CONFIGURAÇÕES E MAPEAMENTO
# ==========================================================

COLUNAS_CHAVE = {
    'COD_PRODUTO': 'Produto',
    'REFERENCIA': 'Referência',
    'DESCRICAO': 'Descrição Produto',
    'CATEGORIA': 'GRUPO',
    'SUBGRUPO': 'SUBGRUPO',
    'ESTOQUE': 'Estoque'
}

# Caminho atualizado conforme sua solicitação
DIRETORIO_BASE = os.path.abspath(r'C:\Users\lucasbarros\OneDrive - CTC FRANCHISING S A\Área de Trabalho\Scripts Python')
ARQUIVO_BASE_SANKHYA = os.path.join(DIRETORIO_BASE, 'Base Temporária de Imagens.xlsx')
DIRETORIO_IMAGENS = os.path.join(DIRETORIO_BASE, 'IMAGENS')

SUBGRUPOS_EXCLUIDOS = ['CHINELO', 'SACOLAS', 'REVENDA', 'MARY KAY']

# ==========================================================
# 🚀 GERADOR DE CATÁLOGO COMPLETO
# ==========================================================

def gerar_catalogo():
    try:
        # Lendo a planilha específica
        df_bruto = pd.read_excel(ARQUIVO_BASE_SANKHYA, sheet_name='BASE PRODUTOS')
        print(f"Arquivo '{os.path.basename(ARQUIVO_BASE_SANKHYA)}' lido. Total: {len(df_bruto)} linhas.")
    except Exception as e:
        print(f"❌ ERRO ao ler planilha: {e}")
        return

    df_bruto.columns = df_bruto.columns.str.strip()
    df_filtrado = df_bruto.copy()

    # 1) Exclusão de subgrupos
    if SUBGRUPOS_EXCLUIDOS:
        excluir = [s.upper().strip() for s in SUBGRUPOS_EXCLUIDOS]
        df_filtrado = df_filtrado[~df_filtrado[COLUNAS_CHAVE['SUBGRUPO']].astype(str).str.upper().str.strip().isin(excluir)]

    # 2) Filtro de estoque opcional
    opcao_estoque = input("\nDeseja ocultar produtos sem estoque? (S/N): ").strip().upper()
    if opcao_estoque == "S" and COLUNAS_CHAVE['ESTOQUE'] in df_filtrado.columns:
        df_filtrado = df_filtrado[df_filtrado[COLUNAS_CHAVE['ESTOQUE']] > 0]

    if df_filtrado.empty:
        print("⚠️ Nenhum dado para processar.")
        return

    # Exportação
    nome_saida = f"Catalogo_Vinculado_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho_saida = os.path.join(DIRETORIO_BASE, nome_saida)

    # Selecionando apenas as colunas desejadas para o catálogo
    cols_export = [COLUNAS_CHAVE['COD_PRODUTO'], COLUNAS_CHAVE['REFERENCIA'], 
                   COLUNAS_CHAVE['DESCRICAO'], COLUNAS_CHAVE['CATEGORIA'], 
                   COLUNAS_CHAVE['SUBGRUPO']]

    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        df_filtrado[cols_export].to_excel(writer, sheet_name='Catalogo', index=False, startrow=1)

    # 3) Inserção de Imagens via Openpyxl
    wb = load_workbook(caminho_saida)
    ws = wb['Catalogo']
    ws.views.sheetView[0].showGridlines = False
    
    style_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=2, column=6, value="FOTO").alignment = style_center

    # Configuração de layout
    larguras = {'A': 12, 'B': 15, 'C': 40, 'D': 18, 'E': 18, 'F': 20}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    print("→ Vinculando imagens às referências...")
    for index, row in df_filtrado.reset_index(drop=True).iterrows():
        codigo = str(row[COLUNAS_CHAVE['COD_PRODUTO']]).strip()
        row_excel = index + 3 
        ws.row_dimensions[row_excel].height = 90 

        # Busca imagem
        img_path = None
        for ext in ['.jpg', '.png', '.jpeg']:
            p = os.path.join(DIRETORIO_IMAGENS, f"{codigo}{ext}")
            if os.path.exists(p):
                img_path = p
                break

        if img_path:
            try:
                img = OpenpyxlImage(img_path)
                img.width, img.height = 110, 110
                ws.add_image(img, f'F{row_excel}')
            except:
                ws.cell(row=row_excel, column=6, value="Erro Img")
        else:
            ws.cell(row=row_excel, column=6, value="N/A")

        for c in range(1, 6):
            ws.cell(row=row_excel, column=c).alignment = style_center

    wb.save(caminho_saida)
    print(f"🚀 Sucesso! Arquivo gerado: {nome_saida}")

if __name__ == "__main__":
    gerar_catalogo()