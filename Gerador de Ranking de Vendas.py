import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

ARQUIVO_BASE = "Base de Produtos Mais Vendidos.xlsx"
PASTA_IMAGENS = "IMAGENS"
EXCLUIR_SUBGRUPOS = ['CHINELO', 'SACOLAS', 'REVENDA', 'MARY KAY']

def buscar_imagem(codigo):
    for ext in ['jpg', 'png', 'jpeg']:
        caminho = os.path.join(PASTA_IMAGENS, f"{codigo}.{ext}")
        if os.path.exists(caminho):
            return caminho
    return None

def criar_relatorio():
    df = pd.read_excel(ARQUIVO_BASE, sheet_name='BASE PRODUTOS')
    df.columns = df.columns.str.strip()
    
    print("\n--- GERADOR DE TOP 10 ---")
    print("Filtrar por: 1-UF | 2-Região | 3-Inside Sales")
    op = input("Escolha: ")
    
    mapeamento = {"1": "COD. UF", "2": "REGIÃO", "3": "ESTADO/IS"}
    coluna_filtro = mapeamento.get(op)
    
    if not coluna_filtro:
        print("Opção inválida!"); return

    valor_filtro = input(f"Informe o {coluna_filtro}: ").upper().strip()
    
    df = df[~df['SUBGRUPO'].str.upper().isin(EXCLUIR_SUBGRUPOS)]
    
    if valor_filtro:
        df = df[df[coluna_filtro].astype(str).str.upper() == valor_filtro]

    if input("Excluir sem estoque? (S/N): ").upper() == 'S':
        df = df[df['Estoque'] > 0]

    top10 = (df.groupby(['Cód. Produto', 'Referência', 'Descrição Produto', 'GRUPO'])['Quantidade']
             .sum()
             .reset_index()
             .sort_values('Quantidade', ascending=False)
             .head(10))

    nome_saida = f"Relatorio_Top10_{valor_filtro or 'GERAL'}.xlsx"
    top10.to_excel(nome_saida, index=False, startrow=1)
    
    wb = load_workbook(nome_saida)
    ws = wb.active
    ws.insert_cols(3)
    ws.cell(row=2, column=3, value="FOTO")

    for i, cod in enumerate(top10['Cód. Produto'], start=3):
        ws.row_dimensions[i].height = 60
        ws.column_dimensions['C'].width = 15
        
        caminho_img = buscar_imagem(str(cod))
        if caminho_img:
            img = ExcelImage(caminho_img)
            img.width, img.height = 70, 70
            ws.add_image(img, f'C{i}')
        else:
            ws.cell(row=i, column=3, value="Sem Foto")

    wb.save(nome_saida)
    print(f"✅ Arquivo gerado: {nome_saida}")

if __name__ == "__main__":
    criar_relatorio()