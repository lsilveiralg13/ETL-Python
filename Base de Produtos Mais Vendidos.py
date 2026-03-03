import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Font, Alignment
import time

# ==========================================================
# 🔧 CONFIGURAÇÕES
# ==========================================================

COLUNAS_CHAVE = {
    'COD_PRODUTO': 'Cód. Produto',
    'REFERENCIA': 'Referência',
    'DESCRICAO': 'Descrição Produto',
    'CATEGORIA': 'GRUPO',
    'SUBGRUPO': 'SUBGRUPO',
    'COD_UF': 'COD. UF',
    'QUANTIDADE': 'Quantidade',
    'VALOR': 'Faturado',
    'ESTOQUE': 'Estoque',
    'ESTADO_IS': 'ESTADO/IS',
    'REGIAO': 'REGIÃO',
}

DIRETORIO_BASE = os.path.abspath(r'C:\Users\lucasbarros\OneDrive - CTC FRANCHISING S A\Área de Trabalho\Scripts Python')
ARQUIVO_BASE_SANKHYA = os.path.join(DIRETORIO_BASE, 'Base de Produtos Mais Vendidos.xlsx')
DIRETORIO_IMAGENS = os.path.join(DIRETORIO_BASE, 'IMAGENS')

# Lista de exclusão rigorosa
SUBGRUPOS_EXCLUIDOS = ['CHINELO', 'SACOLAS', 'REVENDA', 'MARY KAY']

# ==========================================================
# 🚀 FUNÇÃO PRINCIPAL
# ==========================================================

def consolidar_relatorio(filtros: dict, top_n: int):
    try:
        df_bruto = pd.read_excel(ARQUIVO_BASE_SANKHYA, sheet_name='BASE PRODUTOS')
        print(f"\n[DEBUG] Base carregada: {len(df_bruto)} linhas.")
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return

    # Limpeza de cabeçalho
    df_bruto.columns = df_bruto.columns.str.strip()

    # Mapeamento de colunas
    mapa = {v: k for k, v in COLUNAS_CHAVE.items()}
    df_bruto.rename(columns=mapa, inplace=True)

    # --- BLOCO DE EXCLUSÃO CRÍTICO ---
    for col_ref in ['SUBGRUPO', 'DESCRICAO', 'CATEGORIA']:
        if col_ref in df_bruto.columns:
            df_bruto[col_ref] = df_bruto[col_ref].astype(str).str.upper().str.strip()

    antes_excl = len(df_bruto)
    termos_proibidos = [str(t).upper().strip() for t in SUBGRUPOS_EXCLUIDOS]
    
    for termo in termos_proibidos:
        df_bruto = df_bruto[~df_bruto['SUBGRUPO'].str.contains(termo, na=False)]
        df_bruto = df_bruto[~df_bruto['DESCRICAO'].str.contains(termo, na=False)]
    
    print(f"[DEBUG] Linhas removidas pela lista negra ({', '.join(termos_proibidos)}): {antes_excl - len(df_bruto)}")

    # --- FILTROS DE USUÁRIO ---
    df_filtrado = df_bruto.copy()
    for col, val in filtros.items():
        if val and col in df_filtrado.columns:
            v_upper = str(val).upper().strip()
            df_filtrado = df_filtrado[df_filtrado[col].astype(str).str.upper() == v_upper]
            print(f"[DEBUG] Filtro {col}={v_upper}: Restam {len(df_filtrado)} linhas.")

    # Filtro de Estoque
    if input("\nExcluir sem estoque? (S/N): ").strip().upper() == "S":
        df_filtrado = df_filtrado[pd.to_numeric(df_filtrado['ESTOQUE'], errors='coerce').fillna(0) > 0]
        print(f"[DEBUG] Após filtro de estoque: {len(df_filtrado)} linhas.")

    # --- AJUSTE SOLICITADO: RANKING E CORTE EXATO ---
    df_agrup = df_filtrado.groupby(['REFERENCIA', 'COD_PRODUTO', 'DESCRICAO', 'CATEGORIA']).agg({'QUANTIDADE': 'sum'}).reset_index()
    
    # Ordena por quantidade descrescente e usa o código como desempate para garantir o corte exato
    df_final = (
        df_agrup.sort_values(by=['CATEGORIA', 'QUANTIDADE', 'COD_PRODUTO'], ascending=[True, False, True])
        .groupby('CATEGORIA')
        .head(top_n)
        .reset_index(drop=True)
    )

    # Cria o ranking sequencial limpo
    df_final['RANKING'] = df_final.groupby('CATEGORIA').cumcount() + 1
    # ------------------------------------------------

    print(f"\n✅ Relatório processado com Top {top_n}. Itens finais: {len(df_final)}")
    
    if len(df_final) > 0:
        exportar_excel(df_final, filtros, top_n)
    else:
        print("⚠️ Nenhum dado restou após os filtros.")

def exportar_excel(df, filtros, top_n):
    uf = filtros.get('COD_UF') or filtros.get('REGIAO') or filtros.get('ESTADO_IS') or 'GERAL'
    cat = filtros.get('CATEGORIA', 'GERAL')
    nome_arq = f"Relatorio_Top{top_n}_{uf}_{cat}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho = os.path.join(DIRETORIO_BASE, nome_arq)

    df_f = df[['REFERENCIA', 'COD_PRODUTO', 'DESCRICAO', 'RANKING', 'QUANTIDADE']].copy()
    df_f.columns = ['Rótulos de Linha', 'Cod. Sankhya', 'Descrição', 'Ranking', 'VENDAS_TOTAIS']

    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df_f.to_excel(writer, sheet_name='Relatorio', index=False, startrow=1)

    wb = load_workbook(caminho)
    ws = wb['Relatorio']
    ws.insert_cols(4)
    ws.cell(row=2, column=4, value="Foto")

    # Ajuste de Dimensões
    larguras = {'A': 18, 'B': 15, 'C': 45, 'D': 22, 'E': 12, 'F': 15}
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30

    # Título estilizado
    ws.merge_cells('A1:F1')
    ws['A1'] = f"TOP {top_n} - {cat} - {uf}"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Inserir Imagens e Formatar Células
    for i, row in df.reset_index(drop=True).iterrows():
        r_idx = i + 3
        ws.row_dimensions[r_idx].height = 60 
        
        for col_num in range(1, 7):
            ws.cell(row=r_idx, column=col_num).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cod = str(row['COD_PRODUTO']).strip()
        for ext in ['.jpg', '.png', '.jpeg']:
            p = os.path.join(DIRETORIO_IMAGENS, f"{cod}{ext}")
            if os.path.exists(p):
                img = OpenpyxlImage(p)
                img.width, img.height = 100, 100 
                ws.add_image(img, f'D{r_idx}')
                break
    
    wb.save(caminho)
    print(f"🚀 Sucesso! Arquivo gerado: {nome_arq}")

if __name__ == "__main__":
    print("\n1-UF | 2-Região | 3-Inside Sales")
    op = input("Opção: ")
    top = int(input("Deseja Top 10, 20, 30 ou 40? "))
    
    f = {'CATEGORIA': input("Categoria: ")}
    if op == "1": f['COD_UF'] = input("UF: ")
    elif op == "2": f['REGIAO'] = input("Região: ")
    elif op == "3": f['ESTADO_IS'] = input("Vendedora: ")
    
    consolidar_relatorio(f, top)