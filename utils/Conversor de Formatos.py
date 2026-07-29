import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def formatar_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    fonte_corpo = Font(name='Calibri', size=9)
    fonte_cabecalho = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
    
    alinhamento_central = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    
    preenchimento_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Formatação das linhas de dados
    for row in ws.iter_rows(min_row=2):  # Pula o cabeçalho inicialmente
        for cell in row:
            cell.font = fonte_corpo
            
            # Ajuste de alinhamento inteligente para textos longos (ex: nomes de produtos)
            if cell.value and len(str(cell.value)) > 15:
                cell.alignment = alinhamento_esquerda
            else:
                cell.alignment = alinhamento_central

    # Formatação do cabeçalho (Linha 1)
    for cell in ws[1]:
        cell.font = fonte_cabecalho
        cell.fill = preenchimento_preto
        cell.alignment = alinhamento_central
    
    # Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        # Limita a largura máxima para colunas gigantescas não quebrarem o layout
        ws.column_dimensions[column].width = min(max_length + 3, 50)

    wb.save(caminho_arquivo)


def ler_csv_com_varredura(caminho):
    """Tenta ler o CSV testando os encodings sugeridos."""
    encodings = [
        'utf-8-sig',    # UTF-8 com BOM (Evita quebras no Excel)
        'utf-8',        # UTF-8 padrão
        'cp1252',       # ANSI / Windows-1252
        'iso-8859-1',   # Latin-1
        'utf-16'
    ]
    
    for encoding in encodings:
        try:
            df = pd.read_csv(caminho, encoding=encoding, sep=None, engine='python')
            print(f"✔️ Sucesso ao ler com encoding: {encoding}")
            return df
        except (UnicodeDecodeError, Exception):
            continue
    
    raise Exception("Não foi possível converter o arquivo com nenhum dos encodings da lista.")


def tratar_erros_e_caracteres(df):
    """
    Varre o DataFrame aplicando uma biblioteca robusta de padrões de acentuação corrompidos
    e correções ortográficas universais para o ecossistema Excel/CSV.
    """
    print("\nIniciando tratamento de caracteres e acentuação com biblioteca robusta...")
    
    # BIBLIOTECA EXPANDIDA: Mapeia padrões de quebras de codificação (Mojibake) e padrões gramaticais
    biblioteca_acentuacao = {
        # Padrões comuns de interrogações duplas ou caracteres corrompidos em sistemas legados
        r'\?\?': 'ç',  # Caso geral de caracteres especiais perdidos
        r'([aA])\?([oO])': r'\1ão', # Captura 'a?o' -> 'ão' (ex: rota??o, rota?o, inclina?o)
        r'([aA])\?([asAS])': r'\1ãs', # Captura 'a?as' -> 'ãs'
        r'([eE])\?([mI])': r'\1ê', # Captura 'e?m' -> 'êm' ou 'e?i' -> 'êi'
        
        # Correções cirúrgicas de palavras-chave de inventário, produção e logística
        r'[eE]rgon\?[mI]nico': 'Ergonômico',
        r'ergon\?[mI]nico': 'ergonômico',
        r'[iI]nforma\?\?o': 'Informação',
        r'[iI]nforma\?o': 'Informação',
        r'[pP]re\?\?o': 'Preço',
        r'[pP]re\?o': 'Preço',
        
        # Padrões comuns de exportações de ERPs corrompidas (UTF-8 lido como Windows-1252)
        'Ã§Ã£o': 'ção', 'Ã£o': 'ão', 'Ã§': 'ç', 'Ã¡': 'á', 'Ã©': 'é', 
        'Ã\xad': 'í', 'Ã³': 'ó', 'Ãº': 'ú', 'Ãª': 'ê', 'Ã´': 'ô',
        'Ã\x81': 'Á', 'Ã\x89': 'É', 'Ã\x8d': 'Í', 'Ã\x93': 'Ó', 'Ã\x9a': 'Ú',
        'Ã\x87': 'Ç', 'Ã\x83': 'Ã', 'Â°': '°', 'Âº': 'º', 'Âª': 'ª'
    }

    for col in df.columns:
        # Aplicar correções apenas em colunas que contêm texto
        if df[col].dtype == 'object':
            
            # 1. Correção de Encoding em nível de byte (se a string veio como ISO pura ou CP1252)
            try:
                df[col] = df[col].apply(lambda x: x.encode('cp1252').decode('utf-8') if isinstance(x, str) else x)
            except:
                try:
                    df[col] = df[col].apply(lambda x: x.encode('iso-8859-1').decode('utf-8') if isinstance(x, str) else x)
                except:
                    pass
            
            # Converte para string para aplicar as substituições em lote
            df[col] = df[col].astype(str)

            # 2. Varredura da nossa biblioteca usando Regex (Expressões Regulares)
            # Isso mata variações como 'rotação', 'inclinação', 'combinação' usando uma regra só
            for padrao, correto in biblioteca_acentuacao.items():
                df[col] = df[col].str.replace(padrao, correto, regex=True)
                
            # 3. Limpeza de espaços em branco nas pontas e remoção de strings de erro 'nan'
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
            df[col] = df[col].replace('nan', None) # Corrige efeito colateral do astype(str) em nulos
            
            # 4. Inteligência de Data (Converte colunas de data para o formato real)
            if 'data' in col.lower() or 'dt_' in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors='ignore', dayfirst=True)
                except:
                    pass

    print("✔️ Varredura da biblioteca de acentuação e limpeza concluídas!")
    return df


def converter_arquivos():
    diretorio = r'C:\Users\lucas.barros\OneDrive - BELMICRO TECNOLOGIA SA\Área de Trabalho\Scripts Python'
    
    if not os.path.exists(diretorio):
        print(f"Erro: O caminho não foi encontrado.")
        return

    arquivos_disponiveis = [f for f in os.listdir(diretorio) if os.path.isfile(os.path.join(diretorio, f))]
    
    print(f"\n--- Arquivos em: {diretorio} ---")
    for i, arq in enumerate(arquivos_disponiveis):
        print(f"[{i}] {arq}")

    try:
        idx = int(input("\nDigite o número do arquivo de ORIGEM: "))
        nome_arquivo_entrada = arquivos_disponiveis[idx]
    except Exception as e:
        print(f"Seleção inválida. Erro: {e}")
        return

    ext_saida = input("Qual a extensão de SAÍDA desejada (.xlsx, .parquet, .csv): ").strip().lower()
    if not ext_saida.startswith('.'): ext_saida = '.' + ext_saida

    caminho_entrada = os.path.join(diretorio, nome_arquivo_entrada)
    nome_base = os.path.splitext(nome_arquivo_entrada)[0]
    caminho_saida = os.path.join(diretorio, nome_base + ext_saida)

    print(f"\nProcessando: {nome_arquivo_entrada}...")
    
    try:
        # --- 1. LEITURA DOS DADOS ---
        if nome_arquivo_entrada.lower().endswith('.csv'):
            df = ler_csv_com_varredura(caminho_entrada)
        elif nome_arquivo_entrada.lower().endswith('.parquet'):
            df = pd.read_parquet(caminho_entrada)
        elif nome_arquivo_entrada.lower().endswith(('.xlsx', '.xlsm')):
            df = pd.read_excel(caminho_entrada)
        else:
            print("Extensão não suportada.")
            return

        # --- 2. TRATAMENTO INTERMEDIÁRIO ---
        df = tratar_erros_e_caracteres(df)

        # --- 3. EXPORTAÇÃO E GRAVAÇÃO FINAL ---
        if ext_saida in ['.xlsx', '.xlsm']:
            df.to_excel(caminho_saida, index=False)
            formatar_excel(caminho_saida)
        elif ext_saida == '.parquet':
            df.to_parquet(caminho_saida)
        elif ext_saida == '.csv':
            df.to_csv(caminho_saida, index=False, encoding='utf-8-sig')
        
        print(f"\n✅ SUCESSO! Salvo como: {os.path.basename(caminho_saida)}")

    except Exception as e:
        print(f"\n❌ Erro crítico: {e}")

if __name__ == "__main__":
    converter_arquivos()