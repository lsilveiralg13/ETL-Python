import pandas as pd
from datetime import datetime, timedelta
import numpy as np
import openpyxl
import shutil
import os

# 1. DEFINIÇÕES GLOBAIS
NOME_ARQUIVO_ENTRADA = '9 - CRM MM - Set.25.xlsx'
NOMES_ABAS = ['ANDRÉ', 'DUDA', 'RAIANE', 'ROBERTA', 'SCARLAT', 'TARVYLLA']

# Colunas preenchidas MANUALMENTE que precisam de limpeza
COLUNAS_MANUAIS = [
    'Nome do Cliente', 'Contato', 'CNPJ', 'Instagram', 
    'Status do Lead', 'Fonte do Lead', 'Status Conclusão', 
    'Cidade' 
]

# Colunas de data que contêm fórmulas (Ex: HOJE()) e precisam ser "congeladas"
COLUNAS_DATA_CONGELAR = [
    'DATA',                 # Coluna 'C'
    'Data do Próximo Contato', # Coluna 'U'
    'Dias desde Últim. Contato' # Coluna 'W'
]

# --- FUNÇÕES AUXILIARES (Manter inalteradas se não forem a causa do erro) ---

def aplicar_correcoes_e_limpeza(df):
    """Aplica regras de correção e padronização em um DataFrame (aba única)."""
    
    # Normaliza os nomes das colunas para facilitar o acesso
    df.columns = df.columns.str.upper().str.strip() 

    # --- A. LIMPEZA E PADRONIZAÇÃO DE TEXTO NAS COLUNAS MANUAIS ---
    for col in COLUNAS_MANUAIS:
        col_upper = col.upper().strip()
        if col_upper in df.columns:
            # Converte para string, remove espaços e padroniza para MAIÚSCULAS
            df[col_upper] = df[col_upper].astype(str).str.strip().str.upper()
            
            # Substitui valores "vazios" digitados por NaN
            df[col_upper] = df[col_upper].replace({'NAN', 'NA', 'N/A', '-'}, np.nan)
        
    # --- B. CORREÇÕES ESPECÍFICAS DE PICKLISTS (Exemplo) ---
    if 'STATUS DO LEAD' in df.columns:
        df['STATUS DO LEAD'] = df['STATUS DO LEAD'].replace({
            'FECHADO': 'CONVERTIDO',
            'PERDI': 'PERDIDO',
            'LIGAR DEPOIS': 'FOLLOW-UP'
        })
        
    return df

def congelar_datas(df_limpo, nome_coluna_data):
    """Aplica a lógica de congelamento, substituindo a data de hoje pela de ontem."""
    
    data_ontem = (datetime.now() - timedelta(days=1)).date()
    data_hoje = datetime.now().date()
    
    # Converte a coluna para data, ignorando erros
    data_formatada = pd.to_datetime(df_limpo[nome_coluna_data], errors='coerce').dt.date

    for idx, data_celula in data_formatada.items():
        if data_celula == data_hoje:
            df_limpo.loc[idx, nome_coluna_data] = data_ontem
        elif pd.isna(data_celula):
             df_limpo.loc[idx, nome_coluna_data] = None

    return df_limpo


# --- FUNÇÃO PRINCIPAL CORRIGIDA ---

def processar_crm_com_formatacao(nome_arquivo):
    """Lê o arquivo, processa todas as abas, sobrescreve dados mantendo a formatação."""
    
    # CUIDADO: O código irá criar um backup e depois SOBRESCREVER o arquivo original.
    data_backup = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_backup = f'CRM_BACKUP_{data_backup}.xlsx'

    try:
        # 0. Criar backup do arquivo original
        shutil.copy(nome_arquivo, nome_backup)
        print(f"Backup criado com sucesso: {nome_backup}")

        # 1. Abrir o arquivo original com openpyxl (modo de escrita e leitura)
        # Ignoramos warnings de validação de dados, pois não afetam a escrita
        book = openpyxl.load_workbook(nome_arquivo)
        
        abas_processadas = 0

        # 2. Loop principal para processar cada aba de SDR
        for aba_nome in NOMES_ABAS:
            if aba_nome in book.sheetnames:
                print(f"-> Corrigindo dados na aba: {aba_nome}...")
                
                # --- A. LEITURA E LIMPEZA COM PANDAS ---
                df_aba = pd.read_excel(nome_arquivo, sheet_name=aba_nome)
                df_aba_limpo = aplicar_correcoes_e_limpeza(df_aba.copy())
                
                # Aplica o congelamento de datas
                for col_data in COLUNAS_DATA_CONGELAR:
                    col_upper = col_data.upper().strip()
                    if col_upper in df_aba_limpo.columns:
                        df_aba_limpo = congelar_datas(df_aba_limpo, col_upper)
                
                # --- B. SOBRESCREVER DADOS NA PLANILHA COM OPENPYXL ---
                ws = book[aba_nome]
                
                # CORREÇÃO CRÍTICA: Trata células de cabeçalho nulas (None)
                # Garante que .upper() e .strip() só sejam aplicados a strings válidas.
                header = []
                for cell in ws[1]:
                    value = cell.value
                    if value is not None:
                        header.append(str(value).upper().strip())
                    else:
                        header.append(None) # Mantém a posição se a célula estiver vazia
                
                # Mapeamento para encontrar o índice correto da coluna no Excel (1-based)
                # Filtra apenas colunas com nome válido
                col_map = {name: idx + 1 for idx, name in enumerate(header) if name is not None}

                # Colunas para sobrescrever (todas as colunas do DataFrame limpo)
                cols_to_write = df_aba_limpo.columns
                
                # Itera sobre as linhas do DataFrame corrigido (começa da linha 2 do Excel)
                for r_idx, row_series in df_aba_limpo.iterrows():
                    excel_row = r_idx + 2 # Dados começam na linha 2 do Excel
                    
                    # Itera sobre cada coluna corrigida
                    for col_name in cols_to_write:
                        # Verifica se o nome da coluna corrigida existe no mapeamento do Excel
                        if col_name in col_map:
                            excel_col = col_map[col_name]
                            new_value = row_series[col_name]
                            
                            # Se o valor é NaN (Pandas para "vazio"), escreve None no Excel
                            if pd.isna(new_value):
                                ws.cell(row=excel_row, column=excel_col, value=None)
                            else:
                                ws.cell(row=excel_row, column=excel_col, value=new_value)

                abas_processadas += 1
            else:
                print(f"Aviso: Aba '{aba_nome}' não encontrada no arquivo.")

        # 3. SALVAR O ARQUIVO ORIGINAL (SOBRESCREVENDO COM OS DADOS CORRIGIDOS)
        book.save(nome_arquivo)
        
        print("\n" + "=" * 50)
        print(f"Sucesso! {abas_processadas} abas corrigidas.")
        print("A formatação e a aba 'DASH DADOS' foram preservadas.")
        print(f"O arquivo original '{nome_arquivo}' foi atualizado.")
        print("=" * 50)
        
    except FileNotFoundError:
        print(f"ERRO CRÍTICO: O arquivo '{nome_arquivo}' não foi encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro inesperado: {e}")

# EXECUÇÃO
# processar_crm_com_formatacao(NOME_ARQUIVO_ENTRADA)