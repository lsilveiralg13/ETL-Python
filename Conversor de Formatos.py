import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def formatar_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    fonte_corpo = Font(name='Calibri', size=9)
    fonte_cabecalho = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
    
    alinhamento_central = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
    
    preenchimento_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Formatação das linhas de dados
    for row in ws.iter_rows(min_row=2):  # Pula o cabeçalho inicialmente
        for cell in row:
            cell.font = fonte_corpo
            
            # Ajuste de alinhamento inteligente para textos longos (ex: nomes de produtos)
            if cell.value and len(str(cell.value)) > 15:
                cell.alignment = alinhamento_esquerda
            else:
                cell.alignment = alinhamento_central

    # Formatação do cabeçalho (Linha 1)
    for cell in ws[1]:
        cell.font = fonte_cabecalho
        cell.fill = preenchimento_preto
        cell.alignment = alinhamento_central
    
    # Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        # Limita a largura máxima para colunas gigantescas não quebrarem o layout
        ws.column_dimensions[column].width = min(max_length + 3, 50)

    wb.save(caminho_arquivo)

def ler_csv_com_varredura(caminho):
    """Tenta ler o CSV testando os encodings sugeridos."""
    encodings = [
        'utf-8-sig',    # UTF-8 com BOM (Evita quebras no Excel)
        'utf-8',        # UTF-8 padrão
        'cp1252',       # ANSI / Windows-1252
        'iso-8859-1',   # Latin-1
        'utf-16'
    ]
    
    for encoding in encodings:
        try:
            df = pd.read_csv(caminho, encoding=encoding, sep=None, engine='python')
            print(f"✔️ Sucesso ao ler com encoding: {encoding}")
            return df
        except (UnicodeDecodeError, Exception):
            continue
    
    raise Exception("Não foi possível converter o arquivo com nenhum dos encodings da lista.")

def tratar_erros_e_caracteres(df):
    """
    Varre o DataFrame para corrigir problemas de acentuação (Mojibake),
    remover espaços extras e aplicar correções cirúrgicas em strings corrompidas.
    """
    print("\nIniciando tratamento de caracteres e limpeza dos dados...")
    
    # DICIONÁRIO DE CORREÇÃO CIRÚRGICA: Mapeia os erros exatos que aparecem na sua tela
    substituicoes_comuns = {
        'Ergon?mico': 'Ergonômico',
        'ergon?mico': 'ergonômico',
        'rota??o': 'rotação',
        'inclina??o': 'inclinação',
        'Ajuste de inclina??o': 'Ajuste de inclinação',
        'Informa??o': 'Informação',
        'Pre??o': 'Preço',
        '??': 'ç'  # Substituição genérica para o que sobrou de ponto de interrogação duplo
    }

    for col in df.columns:
        # Aplicar correções apenas em colunas que contêm texto
        if df[col].dtype == 'object':
            
            # 1. Tenta limpar decodificações remanescentes tortas
            try:
                df[col] = df[col].apply(lambda x: x.encode('cp1252').decode('utf-8') if isinstance(x, str) else x)
            except:
                try:
                    df[col] = df[col].apply(lambda x: x.encode('iso-8859-1').decode('utf-8') if isinstance(x, str) else x)
                except:
                    pass
            
            # 2. Aplica as substituições do nosso dicionário para limpar os '?' do texto
            for erro, correto in substituicoes_comuns.items():
                df[col] = df[col].astype(str).str.replace(erro, correto, regex=False)
                
            # 3. Limpeza de espaços em branco nas pontas
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
            
            # 4. Inteligência de Data (Converte colunas de data para o formato real)
            if 'data' in col.lower() or 'dt_' in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors='ignore', dayfirst=True)
                except:
                    pass

    print("✔️ Tratamento de texto e formatação de tipos concluídos!")
    return df

def converter_arquivos():
    diretorio = r'C:\Users\lucas.barros\OneDrive - BELMICRO TECNOLOGIA SA\Área de Trabalho\Scripts Python'
    
    if not os.path.exists(diretorio):
        print(f"Erro: O caminho não foi encontrado.")
        return

    arquivos_disponiveis = [f for f in os.listdir(diretorio) if os.path.isfile(os.path.join(diretorio, f))]
    
    print(f"\n--- Arquivos em: {diretorio} ---")
    for i, arq in enumerate(arquivos_disponiveis):
        print(f"[{i}] {arq}")

    try:
        idx = int(input("\nDigite o número do arquivo de ORIGEM: "))
        nome_arquivo_entrada = arquivos_disponiveis[idx]
    except Exception as e:
        print(f"Seleção inválida. Erro: {e}")
        return

    ext_saida = input("Qual a extensão de SAÍDA desejada (.xlsx, .parquet, .csv): ").strip().lower()
    if not ext_saida.startswith('.'): ext_saida = '.' + ext_saida

    caminho_entrada = os.path.join(diretorio, nome_arquivo_entrada)
    nome_base = os.path.splitext(nome_arquivo_entrada)[0]
    caminho_saida = os.path.join(diretorio, nome_base + ext_saida)

    print(f"\nProcessando: {nome_arquivo_entrada}...")
    
    try:
        # --- 1. LEITURA DOS DADOS ---
        if nome_arquivo_entrada.lower().endswith('.csv'):
            df = ler_csv_com_varredura(caminho_entrada)
        elif nome_arquivo_entrada.lower().endswith('.parquet'):
            df = pd.read_parquet(caminho_entrada)
        elif nome_arquivo_entrada.lower().endswith(('.xlsx', '.xlsm')):
            df = pd.read_excel(caminho_entrada)
        else:
            print("Extensão não suportada.")
            return

        # --- 2. TRATAMENTO INTERMEDIÁRIO ---
        df = tratar_erros_e_caracteres(df)

        # --- 3. EXPORTAÇÃO E GRAVAÇÃO FINAL ---
        if ext_saida in ['.xlsx', '.xlsm']:
            df.to_excel(caminho_saida, index=False)
            formatar_excel(caminho_saida)
        elif ext_saida == '.parquet':
            df.to_parquet(caminho_saida)
        elif ext_saida == '.csv':
            df.to_csv(caminho_saida, index=False, encoding='utf-8-sig')
        
        print(f"\n✅ SUCESSO! Salvo como: {os.path.basename(caminho_saida)}")

    except Exception as e:
        print(f"\n❌ Erro crítico: {e}")

if __name__ == "__main__":
    converter_arquivos()